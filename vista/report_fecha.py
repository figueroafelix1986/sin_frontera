import customtkinter as ctk
from tkinter import ttk
from tkcalendar import DateEntry
from datetime import datetime, date
from .list_combo import Lista_combobox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import tkinter.messagebox as messagebox
import pandas as pd
from tkinter import filedialog


class ReporteFecha(ctk.CTkToplevel):
    def __init__(self,controller_registro,controller_hijosmenores):
        super().__init__()
        self.contr_list_comb = Lista_combobox()
        
        self.protocol("WM_DELETE_WINDOW", lambda: self.cerrar_ventana(self))
      
        self.controller_registro=controller_registro
        self.controller_hijosmenores=controller_hijosmenores
        self.title("Reporte por fecha")
        
        self.label_frame_supe = ctk.CTkFrame(self, corner_radius=10, border_width=5)
        self.label_frame_supe.pack(side="top",expand=False,fill="x")
        
        self.label_frame_bajo = ctk.CTkFrame(self, corner_radius=10, border_width=10)
        self.label_frame_bajo.pack(side="top",expand=True,fill="both")
        
         # Configura el peso de las columnas para centrar los DateEntry
        self.label_frame_supe.grid_columnconfigure(0, weight=1)  # Columna vacía a la izquierda
        self.label_frame_supe.grid_columnconfigure(1, weight=0)  # Columna del primer DateEntry
        self.label_frame_supe.grid_columnconfigure(2, weight=0)  # Columna del segundo DateEntry
        self.label_frame_supe.grid_columnconfigure(3, weight=1)  # Columna vacía a la derecha

        
        self.entry_fecha_inic= DateEntry(self.label_frame_supe, date_pattern='dd/mm/yyyy')
        self.entry_fecha_inic.grid(row=0, column=1,  padx=20, pady=10)
        
        self.entry_fecha_fin = DateEntry(self.label_frame_supe, date_pattern='dd/mm/yyyy')
        self.entry_fecha_fin.grid(row=0, column=2,  padx=20, pady=10)   
        
        self.button_aceptar = ctk.CTkButton(self.label_frame_supe, text="Aceptar", command=self.buscar_reporte)
        self.button_aceptar.grid(row=1, column=1,  padx=20, pady=10)         
        
        self.button_exportar = ctk.CTkButton(self.label_frame_supe, 
                                            fg_color="green", 
                                            hover_color="grey",
                                            text_color="white", 
                                            text="Exportar", command=self.exportar_reporte)
        self.button_exportar.grid(row=1, column=2,  padx=20, pady=10) 
        
        
       # Crear el Treeview
        self.tree = ttk.Treeview(self.label_frame_bajo, columns=("ID","Data Cadastro","Cadastrante", "Nombre", "CPF", 
                                                                 "Telefono","Email","Genero","Color de Piel","Etnia",
                                                                 "LGBTQIAPN+","Deficiência","Fech_Nacimiento", "Filhos Menores",
                                                                 "País de Nacionalidade","Estado Civil","Escolaridade",
                                                                 "Data de Chegada ao Brasil","Status Migratório","Situação Laboral",
                                                                 "Atividade econômica","Possui Carteira de Trabalho?",
                                                                 "Benefício do governo?","UF de Residência do Migrante",
                                                                 "Cidade de Residência do Migrante","ENDEREÇO",                                                                 
                                                                 "Renda Familiar Mensal","Ciudad de  Nacimiento","Hijos/Edad/ y Sexo",
                                                                 "Profesión y oficio","Tipo de Solicitud","Observaciones"), show='headings')


        self.tree.column("ID", width=0)  # Ocultar la columna ID
        self.tree.heading("ID", text="")

        # Crear las barras de desplazamiento
        self.scrollbar_y = ttk.Scrollbar(self.label_frame_bajo, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar_y.set)

        self.scrollbar_x = ttk.Scrollbar(self.label_frame_bajo, orient="horizontal", command=self.tree.xview)
        self.tree.configure(xscrollcommand=self.scrollbar_x.set)

        # Empaquetar las barras de desplazamiento
        self.scrollbar_y.pack(side="right", fill='y')
        self.scrollbar_x.pack(side="bottom", fill='x')

        # Empaquetar el Treeview
        self.tree.pack(side="left", expand=True, fill='both')

        # Configuración de encabezados y columnas
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            #self.tree.column(col, width=100)  # Ajusta el ancho según sea necesario
        self.button_exportar.configure(state='disabled')
            

    def buscar_reporte(self):     
        for item in self.tree.get_children():
            self.tree.delete(item)          
        fecha_iniio = self.entry_fecha_inic.get()
        fecha_inicio = datetime.strptime(fecha_iniio, '%d/%m/%Y')
        fecha_inicio = fecha_inicio.replace(hour=0, minute=0, second=0)
        fecha_fin= self.entry_fecha_fin.get()
        fecha_fininal = datetime.strptime(fecha_fin, '%d/%m/%Y')  
        fecha_fininal = fecha_fininal.replace(hour=23, minute=59, second=59)      
        list_reporte=(self.controller_registro.report_fecha(fecha_inicio,fecha_fininal))
        for reporte in list_reporte:                      
            fecha_nacimiento=self.contr_list_comb.format_fecha(reporte.persona.fecha_nacimiento)
            lgbtqiapn=self.contr_list_comb.retorn_opcion(reporte.lgbtqiapn)
            fecha_llegada=self.contr_list_comb.format_fecha(reporte.persona.fecha_llegada)
            beneficio_gobierno=self.contr_list_comb.retorn_opcion(reporte.beneficio_gobierno)
            cartera_trabajo=self.contr_list_comb.retorn_opcion(reporte.cartera_trabajo)
            nino_menores=self.controller_hijosmenores.listar_hijosmenores(reporte.persona_id)
            data_cadastro=self.contr_list_comb.format_fecha(reporte.fecha_creacion)
            # Inicializamos un diccionario para contar géneros y edades
           
            edad_genero= ""
            for menores in nino_menores:
                genero=(menores.genero)
                fecha_nacimiento_hijos=(menores.fecha_nacimiento)
                
               
                edad = (date.today() - fecha_nacimiento_hijos).days // 365  # Edad en años

                # Definimos la clave según el género
                clave_genero = 'h' if genero.lower() == 'feminino' else 'v'
                
                 # Agregamos la información a resultado_ejemplo
                edad_genero += f"{clave_genero} {edad}, "

            if edad_genero:
                edad_genero = edad_genero[:-2]
                
            
            self.tree.insert("", "end", values=(reporte.id,data_cadastro,reporte.cadastrante, reporte.persona.nombre_apellido,reporte.num_ident,
                                               reporte.telefono,reporte.email,reporte.persona.genero.nombre,reporte.persona.raza.nombre,
                                               reporte.etnia,lgbtqiapn,reporte.deficiencia,fecha_nacimiento,reporte.nino_menor,
                                               reporte.persona.pais.nombre,reporte.estado_civil,reporte.persona.nivel_escol.nombre,fecha_llegada,
                                               reporte.status_migrat,reporte.situa_laboral,reporte.activ_econo,cartera_trabajo,
                                               beneficio_gobierno,reporte.uf_residen,reporte.ciudad_resid_brasil,reporte.direccion,reporte.renta_mens,
                                               reporte.persona.ciudad_person,edad_genero,reporte.profec_oficio,reporte.tipo_solicitud,reporte.observacion))

        #self.destroy()
        
        self.button_exportar.configure(state='normal')
        
    def exportar_reporte(self):
    # Obtener los datos del Treeview
        data = []
        for row in self.tree.get_children():
            data.append(self.tree.item(row)['values'])

        # Definir los encabezados
        headers = ["ID","Data Cadastro","Cadastrante", "Nome Completo do Beneficiário", "CPF do Migrante (Apenas Números)", 
                    "Telefone do Beneficiário (apenas números com DDD)","Email do Beneficiário",
                    "Gênero","Raça/Cor","Etnia","LGBTQIAPN+","Deficiência","Data de Nascimento",
                    "Filhos Menores de Idade no Brasil","País de Nacionalidade","Estado Civil",
                    "Escolaridade","Data de Chegada ao Brasil","Status Migratório","Situação Laboral",
                    "Atividade econômica","Possui Carteira de Trabalho?",
                    "Recebe algum benefício do governo?","UF de Residência do Migrante",
                    "Cidade de Residência do Migrante (atenção para ortografia)","ENDEREÇO",                                                                 
                    "Renda Familiar Mensal","Ciudad de  Nacimiento","Hijos/Edad/ y Sexo","Profesión y oficio",
                    "Tipo de Solicitud","Observaciones"]

        # Crear un DataFrame de pandas
        df = pd.DataFrame(data, columns=headers)

        # Abrir un cuadro de diálogo para guardar el archivo
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx"),
                                                            ("All files", "*.*")],
                                                title="Guardar como")

        if file_path:
            # Exportar el DataFrame a un archivo Excel
            df.to_excel(file_path, index=False)
            
            # Abrir el archivo Excel y formatear
            wb = load_workbook(file_path)
            ws = wb.active

            # Ajustar el ancho de las columnas y centrar el texto
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Obtiene la letra de la columna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Añadir un poco de espacio extra
                ws.column_dimensions[column].width = adjusted_width

                # Aplicar ajuste de texto y alineación centrada
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')

            # Guardar los cambios en el archivo
            wb.save(file_path)
            wb.close()
            messagebox.showinfo("Guardado",f"Dato guardado con exito en {file_path}")
            
    def cerrar_ventana(self, main_app):
        # Cierra la ventana
        self.destroy()
        # Muestra la ventana principal de nuevo
        main_app.deiconify()
            